import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Test Credentials'

# Headers
headers = ['Full Name', 'Email Address', 'College Domain', 'Role', 'Password']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='1f7a5d', end_color='1f7a5d', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Test data - one for each role
data = [
    ['Admin User', 'admin@college.edu', 'college.edu', 'COLLEGE_ADMIN', 'CollegeAdmin@123'],
    ['Event Organizer', 'organizer@college.edu', 'college.edu', 'ORGANIZER', 'Organizer@123'],
    ['Event Approver', 'approver@college.edu', 'college.edu', 'APPROVER', 'Approver@123'],
    ['Regular User', 'user@college.edu', 'college.edu', 'USER', 'User@123']
]

# Add data
for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='left', vertical='center')

# Column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 20

# Save
wb.save('test_credentials.xlsx')
print('✅ Test credentials Excel file created successfully')
print('📊 Roles included: COLLEGE_ADMIN, ORGANIZER, APPROVER, USER')
